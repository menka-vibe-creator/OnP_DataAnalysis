"""Excel report generator — creates a formatted .xlsx with data, stats, and charts."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
_TITLE_FONT = Font(bold=True, size=13)


def generate_excel_report(
    csv_path: str | Path,
    analysis_text: str,
    output_path: str | Path,
) -> Path:
    """Create a formatted Excel workbook from CSV data and agent analysis.

    Sheets created:
    - Data      — raw CSV as a formatted table
    - Summary   — descriptive stats + bar charts for each numeric column (up to 4)
    - Analysis  — agent's markdown text

    Args:
        csv_path:      Path to the source CSV file.
        analysis_text: Agent's markdown analysis string.
        output_path:   Destination .xlsx path (parent dirs created if needed).

    Returns:
        The resolved output path.
    """
    csv_path = Path(csv_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(csv_path)
    wb = Workbook()

    _data_sheet(wb.active, df)

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        ws_sum = wb.create_sheet("Summary")
        _summary_sheet(ws_sum, df, numeric_cols)

    ws_txt = wb.create_sheet("Analysis")
    _analysis_sheet(ws_txt, analysis_text)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _data_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "Data"
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif r % 2 == 0:
                cell.fill = _ALT_FILL
    _autowidth(ws)


def _summary_sheet(ws, df: pd.DataFrame, numeric_cols: list[str]) -> None:
    ws.title = "Summary"

    # --- Descriptive statistics table ---
    ws.cell(1, 1, "Descriptive Statistics").font = _TITLE_FONT
    stats = df[numeric_cols].describe().round(2)
    stats_with_index = stats.reset_index()  # 'index' col = stat label
    stats_rows = list(dataframe_to_rows(stats_with_index, index=False, header=True))
    for r, row in enumerate(stats_rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 3:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
    _autowidth(ws)

    stats_end_row = 3 + len(stats_rows)  # last row of stats table
    chart_block_start = stats_end_row + 3  # leave 2 blank rows

    # --- Chart data + bar chart per numeric column (up to 4) ---
    cat_cols = df.select_dtypes(exclude="number").columns.tolist()

    for i, col in enumerate(numeric_cols[:4]):
        block_row = chart_block_start + i * 40  # 40-row vertical spacing

        # Build aggregated series
        if cat_cols:
            group_col = cat_cols[0]
            agg = df.groupby(group_col)[col].sum().reset_index().head(12)
            label_col = group_col
        else:
            agg = pd.DataFrame({
                "Row": range(1, min(len(df), 12) + 1),
                col: df[col].head(12).values,
            })
            label_col = "Row"

        # Write mini data-table (cols A-B)
        ws.cell(block_row, 1, label_col).font = Font(bold=True)
        ws.cell(block_row, 2, col).font = Font(bold=True)
        for j, (_, agg_row) in enumerate(agg.iterrows(), 1):
            ws.cell(block_row + j, 1, str(agg_row[label_col]))
            ws.cell(block_row + j, 2, float(agg_row[col]))

        n = len(agg)

        # Build bar chart referencing the mini table
        chart = BarChart()
        chart.type = "col"
        chart.title = col
        chart.style = 10
        chart.width = 18   # cm
        chart.height = 12  # cm
        chart.legend = None

        data_ref = Reference(ws, min_col=2, min_row=block_row, max_row=block_row + n)
        cats_ref = Reference(ws, min_col=1, min_row=block_row + 1, max_row=block_row + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        anchor_cell = ws.cell(block_row + n + 2, 1).coordinate
        ws.add_chart(chart, anchor_cell)


def _analysis_sheet(ws, text: str) -> None:
    ws.title = "Analysis"
    ws.cell(1, 1, "Agent Analysis").font = _TITLE_FONT
    ws.column_dimensions["A"].width = 120
    for i, line in enumerate(text.splitlines(), 3):
        cell = ws.cell(i, 1, line or " ")
        cell.alignment = Alignment(wrap_text=False)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
